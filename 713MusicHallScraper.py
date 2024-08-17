# You must download the following items in the machine you are running first
# pip install beautifulsoup4
# pip install openpyxl
# pip install datetime
# brew install chromedriver
# pip install selenium
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
from selenium import webdriver


# Venue events page URL
# This part of the code is where you manually type the url where the list of
# events you want to copy are shown on a website. It will save it to a variable
# called "url".
url = "https://www.713musichall.com/shows"

# Set up the webdriver
# This line sets up a Chrome browser environment and save it to a variable
# called "driver"
driver = webdriver.Chrome()

# Navigate to the website
# This opens a Chrome browser window and loads the website in the "url" variable
driver.get(url)

# Scroll to the bottom of the page to trigger lazy loading
# Because many websites use javascript to dynamically load more content on a
# page as you scroll, we needed a way to scroll to the bottom of a page.
# However, using code to scroll directly to the bottom did not work. I had to
# find a way to slowly scroll multiple times to make sure I got to the "real
# bottom of the page.
# This part of the code needs to be fixed to more effectivley load an entire
# JavaScript webpage. Right now I had to manually choose 50 slow scrolls to
# finally get to the bottom of the page with everything loaded.
for i in range(50):  # Scroll 50 times
    driver.execute_script(f"window.scrollTo(0, {i * 100});")  # Scroll 100 pixels at a time
    time.sleep(0.5)  # Wait 0.5 seconds between scrolls

# Get the full HTML content
# This is the pare of the code that retreives all the loaded html text/code and
# saves it to a variable called "html"
html = driver.page_source

# Close the webdriver
driver.quit()

# Parse HTML using BeautifulSoup
# This part of the code take the text we copied into the variable "html" and
# uses BeautifulSoups html.parser function to do work on the html text. Not sure
# exactly what "parsing" does but it save it to a new variable called "soup"
soup = BeautifulSoup(html, 'html.parser')

# Find concert data (assuming it's in a list)
# This part of the code uses the "find_all" function and looks in the "soup"
# variable for the parts of the html that have a 'div' and 'class_" specified
# and saves all those parts of the html code to a new "concerts" variable.
concerts = soup.find_all('div', class_='chakra-linkbox')

# Create an Excel workbook and select the active worksheet
# This starts an Excel sheet where we can save the data we want.
wb = Workbook()
ws = wb.active

# Set header row
# This adds headers to the first row of the excel sheet.
ws['A1'] = 'Date'
ws['B1'] = 'Venue'
ws['C1'] = 'Event'
ws['D1'] = 'Band Names'

# Manually set venue and band names.
# Right now this is just manually saved in the code. It saves it to the
# variables "venue" and "band_names". Future code will pull this from html
venue = '713 Music Hall'
band_names = 'TBD'

# Extract band names and dates, and write to Excel
# This creates a loop where the variable "i" starts at 2 and increases until we
# can no longer find the html code specified. We srart with 2 instead of 1
# because row 1 of the excel sheet has the column headers. So we want to start
# on row 2. The code uses the text saved to the "concerts" variable. It finds
# the instances in the code that have the class_= specified in the parentheses
# and copies the text in that area to the corresponding variables "date" and
# "event_name". It then writes the content save in the variable to the cells
# specified. 
for i, concert in enumerate(concerts, start=2):
    date = concert.find('p', class_='chakra-text css-lfdvoo').text.strip()
    event_name = concert.find('p', class_='chakra-text css-zvlevn').text.strip()
    
    ws[f'A{i}'] = date
    ws[f'B{i}'] = venue
    ws[f'C{i}'] = event_name
    ws[f'D{i}'] = band_names

# Get current timestamp
# This gets the current time stamp and saves it to the variable "timestamps"
# This allows us to have a unique file name for our excel sheet
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Save the Excel file with timestamp
wb.save(f'713 Music Hall Concerts_{timestamp}.xlsx')
